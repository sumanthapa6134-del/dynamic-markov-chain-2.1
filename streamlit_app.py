"""
Markov Chain Q-System Predictor — Streamlit Web Application
============================================================
Rock mass quality forward prediction using Markov Chain TPMs.
"""

import os
import io
import sys
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Resolve TPM path relative to this script (no external folder needed) ───────
_DIR = os.path.dirname(os.path.abspath(__file__))
TPM_FILE = os.path.join(_DIR, "dynamictpm", "transition_probability_matrices.xlsx")

# ── Import core model logic ────────────────────────────────────────────────────
sys.path.insert(0, _DIR)
from markov_predictor import (
    RATINGS, STATE_LABELS, PARAMS, NUM_STATES, RQD_RANGES,
    snap, preprocess, to_states, make_row_vector,
    load_tpms, compute_predictions, interpret, compute_q, classify_q,
)

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS & STYLE
# ══════════════════════════════════════════════════════════════════════════════

PARAM_DESCRIPTIONS = {
    "RQD": "Rock Quality Designation — measures degree of jointing (0–100%)",
    "Jn":  "Joint Set Number — number of joint sets present",
    "Jr":  "Joint Roughness Number — surface roughness of joint walls",
    "Ja":  "Joint Alteration Number — alteration / filling of joints",
    "Jw":  "Joint Water Reduction Factor — water inflow conditions",
    "SRF": "Stress Reduction Factor — stress / fault zone conditions",
}

PARAM_ICONS = {
    "RQD": "🪨", "Jn": "📐", "Jr": "〰️",
    "Ja":  "🧱", "Jw": "💧", "SRF": "⚡",
}

QUALITY_META = {
    "Exceptionally poor": {"color": "#7B0000", "bg": "#FFE8E8", "emoji": "🔴", "range": "Q < 0.01"},
    "Extremely poor":     {"color": "#C0392B", "bg": "#FFECEC", "emoji": "🔴", "range": "0.01–0.1"},
    "Very poor":          {"color": "#E74C3C", "bg": "#FFF0ED", "emoji": "🟠", "range": "0.1–1"},
    "Poor":               {"color": "#E67E22", "bg": "#FFF5E6", "emoji": "🟠", "range": "1–4"},
    "Fair":               {"color": "#D4AC0D", "bg": "#FFFDE6", "emoji": "🟡", "range": "4–10"},
    "Good":               {"color": "#27AE60", "bg": "#EAFAF1", "emoji": "🟢", "range": "10–40"},
    "Very good":          {"color": "#1E8449", "bg": "#D5F5E3", "emoji": "🟢", "range": "40–100"},
    "Extremely good":     {"color": "#1A5276", "bg": "#D6EAF8", "emoji": "🔵", "range": "100–400"},
    "Exceptionally good": {"color": "#154360", "bg": "#D0ECF8", "emoji": "🔵", "range": "Q > 400"},
}

Q_BAND_COLORS = {
    "Exceptionally poor": "#7B0000", "Extremely poor": "#C0392B",
    "Very poor": "#E74C3C",          "Poor": "#E67E22",
    "Fair": "#F1C40F",               "Good": "#2ECC71",
    "Very good": "#1ABC9C",          "Extremely good": "#3498DB",
    "Exceptionally good": "#1A5276",
}

# ── Build state label options for each param (shown in selectbox) ──────────────
STATE_OPTIONS = {}
for p in PARAMS:
    STATE_OPTIONS[p] = [
        f"State {i+1}: {STATE_LABELS[p][i]}"
        for i in range(NUM_STATES[p])
    ]

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Markov Chain Q-System Predictor",
    page_icon="⛰️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom CSS
st.markdown("""
<style>
    /* Main font and background */
    .main { background-color: #f8f9fa; }

    /* Hero banner */
    .hero-banner {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 40%, #0f3460 100%);
        border-radius: 14px;
        padding: 2rem 2.5rem;
        margin-bottom: 1.5rem;
        color: white;
    }
    .hero-title {
        font-size: 2.2rem;
        font-weight: 800;
        letter-spacing: -0.5px;
        margin: 0;
    }
    .hero-subtitle {
        font-size: 1rem;
        opacity: 0.75;
        margin-top: 0.4rem;
    }

    /* Section cards */
    .section-card {
        background: white;
        border-radius: 12px;
        padding: 1.4rem 1.6rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        margin-bottom: 1rem;
    }
    .section-title {
        font-size: 1.05rem;
        font-weight: 700;
        color: #1a1a2e;
        border-bottom: 2px solid #e8ecf0;
        padding-bottom: 0.5rem;
        margin-bottom: 1rem;
    }

    /* Quality badge */
    .quality-badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.85rem;
    }

    /* Metric tiles */
    .metric-tile {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        color: white;
    }

    /* Param card */
    .param-card {
        background: #f0f4ff;
        border-left: 4px solid #4361ee;
        border-radius: 6px;
        padding: 0.6rem 1rem;
        margin-bottom: 0.5rem;
        font-size: 0.88rem;
        color: #2d3748;
    }

    /* Input row label */
    .row-label {
        background: #eef2ff;
        border-radius: 6px;
        padding: 4px 10px;
        font-weight: 600;
        font-size: 0.82rem;
        color: #4361ee;
        display: inline-block;
        margin-bottom: 4px;
    }

    /* Streamlit tweaks */
    div[data-testid="stSelectbox"] label { font-weight: 600; }
    div[data-testid="stMetric"] { background: white; border-radius: 10px; padding: 1rem; }
    .stTabs [data-baseweb="tab"] { font-weight: 600; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# CACHED DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner="Loading transition probability matrices…")
def get_tpms():
    return load_tpms(TPM_FILE)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT (rich, multi-sheet)
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _cell_style(ws, row, col, value, bold=False, fill=None, align="center", font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _thin_border()
    if fill:
        c.fill = fill
    return c


def build_excel(pred_ratings, predictions, input_summary):
    """Build a rich, multi-sheet Excel workbook and return as bytes."""
    wb = Workbook()

    HEADER_DARK = "1A1A2E"
    HEADER_MID  = "16213E"
    HEADER_LITE = "0F3460"
    COL_TEXT    = "FFFFFF"
    ALT_ROW     = "F4F6FB"

    # ── Quality colour map for Q_value column ─────────────────────────────────
    def q_fill(q_val):
        if np.isnan(q_val):
            return None
        cls = classify_q(q_val)
        hex_map = {
            "Exceptionally poor": "7B0000", "Extremely poor": "C0392B",
            "Very poor":          "E74C3C", "Poor":           "E67E22",
            "Fair":               "F1C40F", "Good":           "2ECC71",
            "Very good":          "1ABC9C", "Extremely good": "3498DB",
            "Exceptionally good": "1A5276",
        }
        return PatternFill("solid", fgColor=hex_map.get(cls, "FFFFFF"))

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Q Predictions (main results)
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Q Predictions"

    # Title block
    ws1.merge_cells("A1:I1")
    title_cell = ws1["A1"]
    title_cell.value = "Markov Chain Q-System Prediction Results"
    title_cell.font  = Font(bold=True, size=14, color=COL_TEXT)
    title_cell.fill  = _header_fill(HEADER_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:I2")
    sub = ws1["A2"]
    sub.value = f"Input summary:  {input_summary}"
    sub.font  = Font(italic=True, size=9, color="555555")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 18

    # Column headers (row 3)
    headers = ["Step (j)", "RQD pred.", "Jn pred.", "Jr pred.",
               "Ja pred.", "Jw pred.", "SRF pred.", "Q Value", "Rock Mass Quality"]
    for ci, h in enumerate(headers, 1):
        _cell_style(ws1, 3, ci, h, bold=True,
                    fill=_header_fill(HEADER_MID), font_color=COL_TEXT)
    ws1.row_dimensions[3].height = 22

    col_widths = [10, 12, 10, 10, 10, 10, 12, 12, 24]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for row_idx, j in enumerate(range(1, 31), 4):
        alt = row_idx % 2 == 0
        bg  = PatternFill("solid", fgColor=ALT_ROW) if alt else None
        rqd = pred_ratings.get("RQD", {}).get(j, float("nan"))
        jn  = pred_ratings.get("Jn",  {}).get(j, float("nan"))
        jr  = pred_ratings.get("Jr",  {}).get(j, float("nan"))
        ja  = pred_ratings.get("Ja",  {}).get(j, float("nan"))
        jw  = pred_ratings.get("Jw",  {}).get(j, float("nan"))
        srf = pred_ratings.get("SRF", {}).get(j, float("nan"))
        q   = compute_q(pred_ratings, j)
        ql  = classify_q(q) if not np.isnan(q) else ""

        vals = [j, round(rqd,3), round(jn,3), round(jr,3),
                round(ja,3), round(jw,4), round(srf,3),
                round(q,4) if not np.isnan(q) else "N/A", ql]
        for ci, v in enumerate(vals, 1):
            fill = q_fill(q) if ci == 8 and not np.isnan(q) else bg
            fc   = COL_TEXT if ci == 8 and not np.isnan(q) and ql not in ("Fair",) else "000000"
            _cell_style(ws1, row_idx, ci, v, fill=fill, font_color=fc)

    ws1.freeze_panes = "A4"

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Q Classification Guide
    # ─────────────────────────────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Q Classification Guide")
    ws_guide.merge_cells("A1:D1")
    g1 = ws_guide["A1"]
    g1.value = "Barton Q-System Rock Mass Classification"
    g1.font  = Font(bold=True, size=13, color=COL_TEXT)
    g1.fill  = _header_fill(HEADER_DARK)
    g1.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 26

    for ci, h in enumerate(["Q Range", "Classification", "Typical Conditions", "Support Needs"], 1):
        _cell_style(ws_guide, 2, ci, h, bold=True,
                    fill=_header_fill(HEADER_MID), font_color=COL_TEXT)

    guide_data = [
        ("<0.01",   "Exceptionally poor", "Heavy squeezing / flowing",       "Heavy, continuous support"),
        ("0.01–0.1","Extremely poor",     "Very weak, highly stressed",       "Systematic rock bolts + shotcrete"),
        ("0.1–1",   "Very poor",          "Weak / jointed / some squeezing",  "Systematic bolts + wire mesh"),
        ("1–4",     "Poor",               "Fractured / faulted rock",         "Spot bolts + shotcrete"),
        ("4–10",    "Fair",               "Moderately jointed rock",          "Spot bolts where needed"),
        ("10–40",   "Good",               "Competent, moderately jointed",    "Minimal — occasional bolts"),
        ("40–100",  "Very good",          "Massive / few joints",             "Generally unsupported"),
        ("100–400", "Extremely good",     "Near-intact rock",                 "No support required"),
        (">400",    "Exceptionally good", "Massive intact rock",              "No support required"),
    ]
    hex_list = ["7B0000","C0392B","E74C3C","E67E22","F1C40F","2ECC71","1ABC9C","3498DB","1A5276"]
    for ri, (row_d, hx) in enumerate(zip(guide_data, hex_list), 3):
        fc = "FFFFFF" if ri != 7 else "000000"
        for ci, v in enumerate(row_d, 1):
            _cell_style(ws_guide, ri, ci, v,
                        fill=PatternFill("solid", fgColor=hx), font_color=fc)

    for ci, w in enumerate([14, 22, 38, 32], 1):
        ws_guide.column_dimensions[get_column_letter(ci)].width = w
    ws_guide.freeze_panes = "A3"

    # ─────────────────────────────────────────────────────────────────────────
    # Sheets 3–8 — Per-parameter probability evolution
    # ─────────────────────────────────────────────────────────────────────────
    param_colors = {
        "RQD": "1F4E79", "Jn": "375623", "Jr": "843C0C",
        "Ja":  "4B1C6B", "Jw": "154360", "SRF":"7B2D00",
    }

    for param in PARAMS:
        if param not in predictions:
            continue
        ws_p = wb.create_sheet(f"{param} Probabilities")
        n = NUM_STATES[param]
        hx = param_colors.get(param, HEADER_DARK)

        # Title
        ws_p.merge_cells(f"A1:{get_column_letter(n+1)}1")
        tc = ws_p["A1"]
        tc.value = f"{param} — State Probability Evolution (j = 1 to 30)"
        tc.font  = Font(bold=True, size=12, color=COL_TEXT)
        tc.fill  = _header_fill(hx)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_p.row_dimensions[1].height = 24

        # Headers
        _cell_style(ws_p, 2, 1, "Step (j)", bold=True,
                    fill=_header_fill(hx), font_color=COL_TEXT)
        for si in range(1, n+1):
            label = STATE_LABELS[param][si-1][:28]
            _cell_style(ws_p, 2, si+1, f"S{si}: {label}", bold=True,
                        fill=_header_fill(hx), font_color=COL_TEXT)

        # Data
        for ri, j in enumerate(range(1, 31), 3):
            pvec = predictions[param].get(j, [float("nan")] * n)
            alt = ri % 2 == 0
            bg  = PatternFill("solid", fgColor=ALT_ROW) if alt else None
            _cell_style(ws_p, ri, 1, j, bold=True, fill=bg)
            for si in range(n):
                val = float(pvec[si]) if len(pvec) > si else float("nan")
                _cell_style(ws_p, ri, si+2,
                            round(val, 4) if not np.isnan(val) else "N/A",
                            fill=bg)

        # Column widths
        ws_p.column_dimensions["A"].width = 10
        for si in range(1, n+1):
            ws_p.column_dimensions[get_column_letter(si+1)].width = max(
                18, len(STATE_LABELS[param][si-1][:28]) + 4
            )

        # Conditional colour scale on probability columns
        prob_range = f"B3:{get_column_letter(n+1)}32"
        ws_p.conditional_formatting.add(
            prob_range,
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="FFFFFF",
                end_type="num",   end_value=1,     end_color=hx.lstrip("#"),
            )
        )
        ws_p.freeze_panes = "B3"

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet — State Definitions Reference
    # ─────────────────────────────────────────────────────────────────────────
    ws_ref = wb.create_sheet("State Definitions")
    ws_ref.merge_cells("A1:C1")
    rc = ws_ref["A1"]
    rc.value = "Q-System Parameter State Definitions"
    rc.font  = Font(bold=True, size=13, color=COL_TEXT)
    rc.fill  = _header_fill(HEADER_DARK)
    rc.alignment = Alignment(horizontal="center", vertical="center")
    ws_ref.row_dimensions[1].height = 26

    row_r = 2
    for param in PARAMS:
        # Param header
        ws_ref.merge_cells(f"A{row_r}:C{row_r}")
        ph = ws_ref[f"A{row_r}"]
        ph.value = f"{param} — {PARAM_DESCRIPTIONS[param]}"
        ph.font  = Font(bold=True, size=10, color=COL_TEXT)
        ph.fill  = _header_fill(param_colors.get(param, HEADER_MID))
        ph.alignment = Alignment(horizontal="left", vertical="center")
        ws_ref.row_dimensions[row_r].height = 20
        row_r += 1

        for ci, h in enumerate(["State No.", "Rating Value", "Description"], 1):
            _cell_style(ws_ref, row_r, ci, h, bold=True,
                        fill=_header_fill("C0C0C0"), font_color="000000")
        row_r += 1

        for si, (label, rating) in enumerate(
                zip(STATE_LABELS[param], RATINGS[param]), 1):
            alt = si % 2 == 0
            bg = PatternFill("solid", fgColor=ALT_ROW) if alt else None
            _cell_style(ws_ref, row_r, 1, si, fill=bg)
            _cell_style(ws_ref, row_r, 2, rating, fill=bg)
            _cell_style(ws_ref, row_r, 3, label, fill=bg, align="left")
            row_r += 1
        row_r += 1  # blank separator

    ws_ref.column_dimensions["A"].width = 12
    ws_ref.column_dimensions["B"].width = 14
    ws_ref.column_dimensions["C"].width = 55

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PLOTTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def plot_q_trend(pred_ratings):
    steps = list(range(1, 31))
    q_vals = [compute_q(pred_ratings, j) for j in steps]
    qualities = [classify_q(q) if not np.isnan(q) else "" for q in q_vals]
    colors = [Q_BAND_COLORS.get(ql, "#888") for ql in qualities]

    fig = go.Figure()

    # Quality bands as filled shapes
    band_defs = [
        (0.001, 0.01, "Exceptionally poor", "rgba(123,0,0,0.10)"),
        (0.01,  0.1,  "Extremely poor",     "rgba(192,57,43,0.10)"),
        (0.1,   1,    "Very poor",           "rgba(231,76,60,0.10)"),
        (1,     4,    "Poor",                "rgba(230,126,34,0.10)"),
        (4,     10,   "Fair",                "rgba(241,196,15,0.10)"),
        (10,    40,   "Good",                "rgba(46,204,113,0.10)"),
        (40,    100,  "Very good",           "rgba(26,188,156,0.10)"),
        (100,   400,  "Extremely good",      "rgba(52,152,219,0.10)"),
    ]
    q_min = max(0.0001, min(v for v in q_vals if not np.isnan(v)) * 0.5) if any(not np.isnan(v) for v in q_vals) else 0.001
    q_max = max(v for v in q_vals if not np.isnan(v)) * 2 if any(not np.isnan(v) for v in q_vals) else 400

    for lo, hi, label, fill in band_defs:
        if hi < q_min * 0.5 or lo > q_max * 2:
            continue
        fig.add_hrect(
            y0=lo, y1=hi,
            fillcolor=fill, line_width=0,
            annotation_text=label, annotation_position="right",
            annotation=dict(font_size=9, font_color="#555", xanchor="left"),
        )

    # Main line
    fig.add_trace(go.Scatter(
        x=steps, y=q_vals,
        mode="lines+markers",
        line=dict(width=3, color="#1a1a2e"),
        marker=dict(size=8, color=colors, line=dict(width=1.5, color="white"),
                    symbol="circle"),
        name="Q value",
        hovertemplate="<b>Step j=%{x}</b><br>Q = %{y:.4f}<br>Quality: %{text}<extra></extra>",
        text=qualities,
        customdata=qualities,
    ))

    fig.update_layout(
        title=dict(
            text="<b>Predicted Q-Value Evolution</b>  (j = 1 → 30 chainage steps)",
            font=dict(size=15, color="#1a1a2e"), x=0.01,
        ),
        xaxis=dict(
            title="Chainage step (j)", tickmode="linear", dtick=2,
            showgrid=True, gridcolor="#e8ecf0", zeroline=False,
        ),
        yaxis=dict(
            title="Q Value (log scale)", type="log",
            showgrid=True, gridcolor="#e8ecf0",
        ),
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.15),
        height=460,
        margin=dict(l=60, r=160, t=60, b=60),
        hovermode="x unified",
    )
    return fig


def plot_param_predictions(pred_ratings):
    steps = list(range(1, 31))
    colors = ["#4361EE","#3A0CA3","#7209B7","#F72585","#4CC9F0","#2ECC71"]

    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=[f"{PARAM_ICONS[p]}  {p}" for p in PARAMS],
        vertical_spacing=0.10, horizontal_spacing=0.08,
    )
    for i, (param, color) in enumerate(zip(PARAMS, colors)):
        r, c = divmod(i, 2)
        vals = [pred_ratings.get(param, {}).get(j, np.nan) for j in steps]
        fig.add_trace(
            go.Scatter(
                x=steps, y=vals,
                mode="lines+markers",
                line=dict(width=2.5, color=color),
                marker=dict(size=5, color=color),
                name=param,
                showlegend=True,
                hovertemplate=f"<b>{param}</b> j=%{{x}}: %{{y:.3f}}<extra></extra>",
            ),
            row=r+1, col=c+1,
        )

    fig.update_layout(
        title=dict(
            text="<b>Parameter Rating Predictions</b>  (j = 1 → 30)",
            font=dict(size=14, color="#1a1a2e"), x=0.01,
        ),
        height=680,
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=False,
    )
    fig.update_xaxes(showgrid=True, gridcolor="#e8ecf0", dtick=5)
    fig.update_yaxes(showgrid=True, gridcolor="#e8ecf0")
    return fig


def plot_probability_heatmap(predictions, param):
    if param not in predictions:
        return None
    n = NUM_STATES[param]
    steps = sorted(predictions[param].keys())
    z = []
    for j in steps:
        pvec = predictions[param][j]
        z.append([round(float(v), 4) for v in pvec])

    short_labels = [STATE_LABELS[param][i][:30] for i in range(n)]

    fig = go.Figure(go.Heatmap(
        z=z,
        x=short_labels,
        y=[f"j={j}" for j in steps],
        colorscale="Blues",
        zmin=0, zmax=1,
        text=[[f"{v:.3f}" for v in row] for row in z],
        texttemplate="%{text}",
        textfont=dict(size=8),
        hovertemplate="Step %{y}<br>State: %{x}<br>Probability: %{z:.4f}<extra></extra>",
        colorbar=dict(title="Probability", len=0.8),
    ))
    fig.update_layout(
        title=dict(
            text=f"<b>{PARAM_ICONS[param]} {param}</b> — State Probability Evolution",
            font=dict(size=13, color="#1a1a2e"),
        ),
        xaxis=dict(title="State", tickangle=-30, tickfont=dict(size=8)),
        yaxis=dict(title="Chainage step (j)", autorange="reversed",
                   tickfont=dict(size=9)),
        height=520,
        margin=dict(l=60, r=40, t=60, b=120),
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    return fig


def plot_initial_state_dist(row_vectors):
    fig = make_subplots(
        rows=2, cols=3,
        subplot_titles=[f"{PARAM_ICONS[p]} {p}" for p in PARAMS],
        vertical_spacing=0.15, horizontal_spacing=0.08,
    )
    palette = ["#4361EE","#3A0CA3","#7209B7","#F72585","#4CC9F0","#2ECC71"]

    for i, (param, color) in enumerate(zip(PARAMS, palette)):
        r, c = divmod(i, 3)
        rv = row_vectors[param]
        short_lbl = [f"S{si+1}" for si in range(len(rv))]
        fig.add_trace(
            go.Bar(
                x=short_lbl, y=rv,
                marker_color=color, opacity=0.85,
                name=param,
                hovertemplate=f"<b>{param}</b><br>%{{x}}: %{{y:.3f}}<extra></extra>",
                showlegend=False,
            ),
            row=r+1, col=c+1,
        )
    fig.update_layout(
        title=dict(
            text="<b>Initial State Distribution</b>  (from your inputs)",
            font=dict(size=14, color="#1a1a2e"), x=0.01,
        ),
        height=480,
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    fig.update_yaxes(range=[0, 1.05], showgrid=True, gridcolor="#eee")
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════

if "n_rows" not in st.session_state:
    st.session_state.n_rows = 1
if "results" not in st.session_state:
    st.session_state.results = None


# ══════════════════════════════════════════════════════════════════════════════
# MAIN LAYOUT
# ══════════════════════════════════════════════════════════════════════════════

# ── Hero Banner ────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-banner">
  <div class="hero-title">⛰️ Markov Chain Q-System Predictor</div>
  <div class="hero-subtitle">
    Forward rock mass quality prediction using dynamic Transition Probability Matrices &nbsp;|&nbsp;
    Barton Q-System &nbsp;|&nbsp; Up to 30 chainage steps
  </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Prediction Settings")
    st.markdown("---")

    # Number of input observation rows
    n_rows = st.number_input(
        "Number of field observations",
        min_value=1, max_value=10, value=st.session_state.n_rows, step=1,
        help="Add multiple field measurements to build a richer initial state vector.",
    )
    st.session_state.n_rows = n_rows

    st.markdown("---")
    st.markdown("### 📚 Parameter Guide")
    for p in PARAMS:
        st.markdown(
            f"<div class='param-card'><b>{PARAM_ICONS[p]} {p}</b><br>{PARAM_DESCRIPTIONS[p]}</div>",
            unsafe_allow_html=True,
        )
    st.markdown("---")
    st.markdown(
        "<small style='color:#888'>TPMs embedded from <i>dynamictpm/</i>.<br>"
        "Model: Barton (1974) Q-system.<br>"
        "30 forward chainage steps.</small>",
        unsafe_allow_html=True,
    )


# ── Input Section ─────────────────────────────────────────────────────────────
st.markdown("## 📋 Field Observations Input")
st.markdown(
    "Select the state that best matches your field measurement for each parameter. "
    "Add multiple observations to build a weighted initial probability vector."
)

input_rows = []  # list of dicts: {param: state_index (0-based)}

# Dynamic rows
for row_i in range(st.session_state.n_rows):
    with st.expander(f"🔎 Observation #{row_i + 1}", expanded=True):
        cols = st.columns(3)
        row_data = {}
        for pi, param in enumerate(PARAMS):
            with cols[pi % 3]:
                # Default: roughly mid-range state
                default_idx = NUM_STATES[param] // 2
                chosen = st.selectbox(
                    f"{PARAM_ICONS[param]} **{param}**",
                    options=STATE_OPTIONS[param],
                    index=default_idx,
                    key=f"{param}_{row_i}",
                    help=PARAM_DESCRIPTIONS[param],
                )
                # Extract state index
                state_idx = STATE_OPTIONS[param].index(chosen)
                row_data[param] = state_idx
        input_rows.append(row_data)

# ── Run Button ─────────────────────────────────────────────────────────────────
st.markdown("---")
col_btn, col_info = st.columns([1, 3])
with col_btn:
    run_btn = st.button("🚀 Run Prediction", type="primary", use_container_width=True)
with col_info:
    st.info(
        f"**{st.session_state.n_rows} observation(s)** entered. "
        "Predictions will cover **30 forward chainage steps** using embedded TPMs.",
        icon="ℹ️",
    )

# ══════════════════════════════════════════════════════════════════════════════
# PREDICTION PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

if run_btn:
    with st.spinner("⚙️ Loading TPMs & running Markov chain predictions…"):

        # 1. Convert state indices → rating values
        inputs = {}
        for param in PARAMS:
            rating_vals = []
            for row_data in input_rows:
                state_idx = row_data[param]
                rating_vals.append(RATINGS[param][state_idx])
            inputs[param] = rating_vals

        # 2. Snap & convert (already snapped since we use exact RATINGS values)
        snapped = inputs  # already representative values
        states  = to_states(snapped)

        # 3. Build row probability vectors
        row_vectors = {
            param: make_row_vector(states[param], NUM_STATES[param])
            for param in PARAMS
        }

        # 4. Load TPMs (cached)
        tpms = get_tpms()

        # 5. Compute predictions
        predictions = compute_predictions(row_vectors, tpms)

        # 6. Interpret
        pred_ratings = interpret(predictions, snapped)

        # 7. Build input summary string
        parts = []
        for param in PARAMS:
            vals = [f"{RATINGS[param][row[param]]}" for row in input_rows]
            parts.append(f"{param}=[{', '.join(vals)}]")
        input_summary = "  |  ".join(parts)

        # 8. Build Excel
        excel_bytes = build_excel(pred_ratings, predictions, input_summary)

        st.session_state.results = dict(
            pred_ratings=pred_ratings,
            predictions=predictions,
            row_vectors=row_vectors,
            snapped=snapped,
            states=states,
            input_summary=input_summary,
            excel_bytes=excel_bytes,
            input_rows=input_rows,
        )
    st.success("✅ Prediction complete!", icon="✅")

# ══════════════════════════════════════════════════════════════════════════════
# RESULTS DISPLAY
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.results:
    r = st.session_state.results
    pred_ratings = r["pred_ratings"]
    predictions  = r["predictions"]
    row_vectors  = r["row_vectors"]

    st.markdown("---")
    st.markdown("## 📊 Prediction Results")

    # ── Quick-look metrics (j=1, j=5, j=10, j=20, j=30) ──────────────────────
    st.markdown("### Key Step Snapshots")
    milestone_steps = [1, 5, 10, 20, 30]
    mcols = st.columns(len(milestone_steps))
    for col, j in zip(mcols, milestone_steps):
        q = compute_q(pred_ratings, j)
        ql = classify_q(q) if not np.isnan(q) else "N/A"
        meta = QUALITY_META.get(ql, {"color": "#888", "bg": "#f5f5f5", "emoji": "❓"})
        with col:
            st.markdown(
                f"""
                <div style="
                    background:{meta['bg']};
                    border-left:5px solid {meta['color']};
                    border-radius:10px;
                    padding:14px 12px;
                    text-align:center;
                ">
                  <div style="font-size:1.5rem; font-weight:800; color:{meta['color']};">
                    {'N/A' if np.isnan(q) else f'{q:.3f}'}
                  </div>
                  <div style="font-size:0.75rem; color:#555; font-weight:600;">j = {j}</div>
                  <div style="font-size:0.78rem; color:{meta['color']}; font-weight:700; margin-top:4px;">
                    {meta.get('emoji','')} {ql}
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Main tabs ─────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📈 Q-Value Trend",
        "🔢 Parameter Trends",
        "🌡️ Probability Heatmaps",
        "📋 Full Results Table",
        "🏁 Initial State",
    ])

    # ─────── TAB 1: Q-Value Trend ────────────────────────────────────────────
    with tab1:
        st.plotly_chart(plot_q_trend(pred_ratings), use_container_width=True)

        st.markdown("#### Quality Band Legend")
        lcols = st.columns(5)
        for i, (ql, meta) in enumerate(QUALITY_META.items()):
            with lcols[i % 5]:
                st.markdown(
                    f"<div style='background:{meta['bg']};border-left:4px solid {meta['color']};"
                    f"border-radius:6px;padding:6px 10px;margin:3px 0;font-size:0.78rem;'>"
                    f"<b style='color:{meta['color']};'>{meta['emoji']} {ql}</b><br>"
                    f"<span style='color:#777;'>{meta['range']}</span></div>",
                    unsafe_allow_html=True,
                )

    # ─────── TAB 2: Parameter Trends ─────────────────────────────────────────
    with tab2:
        st.plotly_chart(plot_param_predictions(pred_ratings), use_container_width=True)
        st.caption(
            "RQD uses probability-weighted average rating. "
            "All other parameters show the rating of the most probable state (argmax)."
        )

    # ─────── TAB 3: Probability Heatmaps ─────────────────────────────────────
    with tab3:
        st.markdown(
            "Each heatmap shows how the **probability mass** shifts across states "
            "as the model steps forward through future chainages."
        )
        param_sel = st.selectbox(
            "Select parameter to inspect",
            PARAMS,
            format_func=lambda p: f"{PARAM_ICONS[p]} {p} — {PARAM_DESCRIPTIONS[p][:45]}…",
        )
        hm_fig = plot_probability_heatmap(predictions, param_sel)
        if hm_fig:
            st.plotly_chart(hm_fig, use_container_width=True)

        # State label reference
        st.markdown(f"#### {param_sel} State Reference")
        ref_df = pd.DataFrame({
            "State": [f"S{i+1}" for i in range(NUM_STATES[param_sel])],
            "Rating": RATINGS[param_sel],
            "Description": STATE_LABELS[param_sel],
        })
        st.dataframe(ref_df, use_container_width=True, hide_index=True)

    # ─────── TAB 4: Full Results Table ────────────────────────────────────────
    with tab4:
        rows = []
        for j in range(1, 31):
            rqd = pred_ratings.get("RQD", {}).get(j, np.nan)
            jn  = pred_ratings.get("Jn",  {}).get(j, np.nan)
            jr  = pred_ratings.get("Jr",  {}).get(j, np.nan)
            ja  = pred_ratings.get("Ja",  {}).get(j, np.nan)
            jw  = pred_ratings.get("Jw",  {}).get(j, np.nan)
            srf = pred_ratings.get("SRF", {}).get(j, np.nan)
            q   = compute_q(pred_ratings, j)
            ql  = classify_q(q) if not np.isnan(q) else "N/A"
            rows.append({
                "Step j": j,
                "RQD": round(rqd, 3),
                "Jn": round(jn, 3),
                "Jr": round(jr, 3),
                "Ja": round(ja, 3),
                "Jw": round(jw, 4),
                "SRF": round(srf, 3),
                "Q Value": round(q, 4) if not np.isnan(q) else None,
                "Quality": ql,
            })

        results_df = pd.DataFrame(rows)

        def highlight_quality(row):
            ql = row["Quality"]
            meta = QUALITY_META.get(ql, {})
            color = meta.get("bg", "")
            return [f"background-color: {color}" if color else ""] * len(row)

        styled = results_df.style.apply(highlight_quality, axis=1).format({
            "RQD": "{:.3f}", "Jn": "{:.3f}", "Jr": "{:.3f}",
            "Ja": "{:.3f}",  "Jw": "{:.4f}", "SRF": "{:.3f}",
            "Q Value": lambda v: f"{v:.4f}" if v is not None else "N/A",
        })
        st.dataframe(styled, use_container_width=True, height=600, hide_index=True)

    # ─────── TAB 5: Initial State Distribution ────────────────────────────────
    with tab5:
        st.markdown(
            "These bar charts show the **initial state probability distribution** "
            "derived from your field observations. This is the starting vector "
            "multiplied by each TPM to generate predictions."
        )
        st.plotly_chart(plot_initial_state_dist(row_vectors), use_container_width=True)

        # Show input summary table
        st.markdown("#### Input Summary")
        inp_rows_display = []
        for ri, row_data in enumerate(r["input_rows"]):
            disp_row = {"Observation": f"#{ri+1}"}
            for param in PARAMS:
                si = row_data[param]
                disp_row[param] = f"S{si+1}: {RATINGS[param][si]} ({STATE_LABELS[param][si][:25]}…)"
            inp_rows_display.append(disp_row)
        st.dataframe(pd.DataFrame(inp_rows_display), use_container_width=True, hide_index=True)

    # ── Download section ───────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Download Results")
    dl_col1, dl_col2 = st.columns([1, 2])
    with dl_col1:
        st.download_button(
            label="⬇️ Download Excel Report",
            data=r["excel_bytes"],
            file_name="markov_q_prediction.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    with dl_col2:
        st.markdown(
            """
            **Excel report contains:**
            - 📊 **Q Predictions** — full 30-step results table with quality colour coding
            - 🎨 **Q Classification Guide** — Barton quality bands with descriptions
            - 🌡️ **Per-parameter probability sheets** (6 sheets) — state probability at each step
            - 📚 **State Definitions** — complete reference for all parameter states
            """
        )

else:
    # Welcome placeholder
    st.markdown("---")
    st.info(
        "👆 **Configure your field observations in the input section above, "
        "then click 'Run Prediction' to generate results.**",
        icon="⛰️",
    )
    # Show parameter reference
    with st.expander("📚 Q-System Parameter Reference", expanded=False):
        for param in PARAMS:
            st.markdown(f"**{PARAM_ICONS[param]} {param}** — {PARAM_DESCRIPTIONS[param]}")
            ref_df = pd.DataFrame({
                "State": [f"State {i+1}" for i in range(NUM_STATES[param])],
                "Rating": RATINGS[param],
                "Description": STATE_LABELS[param],
            })
            st.dataframe(ref_df, use_container_width=True, hide_index=True)
            st.markdown("")
